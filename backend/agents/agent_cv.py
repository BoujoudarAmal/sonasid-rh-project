import os
from google import genai
from google.genai import types
from pypdf import PdfReader
import docx
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class AgentTraitementCV:
    def __init__(self, cv_path: str):
        self.cv_path = cv_path
        
        # --- CONFIGURATION DE LA CLÉ API GEMINI ---
        # 1. On cherche dans l'environnement Windows
        api_key = os.environ.get("GEMINI_API_KEY")
        
        # 2. Si non trouvée, met ta clé à la place de "AIzaSy..." ci-dessous :
        if not api_key or api_key == "TA_CLE_API_ICI":
            api_key = "AIzaSy..." # <-- COCHE ET METS TA VRAIE CLÉ ICI
            
        try:
            self.client = genai.Client(api_key=api_key)
        except Exception as e:
            print(f"⚠️ Erreur d'initialisation Gemini (Vérifie ta clé API) : {e}")
            self.client = None

    def _extraire_texte_document(self) -> str:
        """Extrait l'intégralité du texte brut du document (PDF, DOCX ou TXT)."""
        texte = ""
        ext = os.path.splitext(self.cv_path)[1].lower().strip()
        
        try:
            if ext == '.pdf':
                reader = PdfReader(self.cv_path)
                for page in reader.pages:
                    content = page.extract_text()
                    if content:
                        texte += content + "\n"
            elif ext in ['.docx', '.doc']:
                doc = docx.Document(self.cv_path)
                for para in doc.paragraphs:
                    texte += para.text + "\n"
            elif ext == '.txt':
                with open(self.cv_path, 'r', encoding='utf-8', errors='ignore') as f:
                    texte = f.read()
        except Exception as e:
            print(f"⚠️ Erreur de lecture du fichier {self.cv_path} : {e}")
            
        return texte.strip()

    def traiter(self) -> dict:
        """Analyse le texte du CV avec Gemini pour retourner un JSON structuré."""
        texte_cv = self._extraire_texte_document()
        
        donnees_defaut = {
            "Nom": "À renseigner", "Prénom": "À renseigner", "Email": "Non détecté",
            "Téléphone": "Non détecté", "Profil": "Impossible d'analyser le contenu.",
            "Compétences": "À évaluer", "Diplômes": "À évaluer"
        }
        
        if not texte_cv:
            return donnees_defaut

        if not self.client:
            print("❌ Erreur : Impossible d'appeler l'API car le client Gemini n'est pas configuré.")
            return donnees_defaut

        consigne_prompt = (
            "Tu es un expert en recrutement RH pour SONASID. Analyse le texte brut suivant extrait d'un CV "
            "et extrait les informations de manière très précise.\n\n"
            f"Texte du CV :\n{texte_cv}\n\n"
            "Tu dois obligatoirement renvoyer un objet JSON correspondant EXACTEMENT à cette structure :\n"
            "{\n"
            '  "Nom": "NOM DU CANDIDAT (En MAJUSCULES)",\n'
            '  "Prénom": "Prénom (Première lettre en majuscule)",\n'
            '  "Email": "adresse email trouvée ou Non détecté",\n'
            '  "Téléphone": "numéro de téléphone marocain ou international trouvé ou Non détecté",\n'
            '  "Profil": "Une courte phrase de résumé du profil professionnel du candidat",\n'
            '  "Compétences": "Liste des compétences clés séparées par des virgules (ex: Python, SQL, Excel)",\n'
            '  "Diplômes": "Le diplôme le plus élevé ou pertinent trouvé (ex: Master, Licence)"\n'
            "}\n"
            "Ne réponds rien d'autre que l'objet JSON brut. Pas de texte avant, pas de texte après."
        )

        try:
            response = self.client.models.generate_content(
                model='gemini-2.5-flash',
                contents=consigne_prompt,
                config=types.GenerateContentConfig(
                    response_mime_type="application/json"
                )
            )
            return json.loads(response.text.strip())
        except Exception as e:
            print(f"⚠️ Erreur lors de l'appel à l'API Gemini : {e}")
            return donnees_defaut

    def calculer_score(self, poste_vise: str) -> int:
        if not poste_vise or poste_vise == "Spontané":
            return 50
        return 85

    @staticmethod
    def exporter_excel(candidats_list: list, chemin_excel: str = None):
        base_dir = os.path.dirname(os.path.abspath(__file__))
        backend_dir = os.path.dirname(base_dir)
        chemin_final = chemin_excel if chemin_excel else os.path.join(backend_dir, "data", "candidats_export.xlsx")
        
        data = []
        for c in candidats_list:
            data.append({
                "Prénom": c.get("Prénom", c.get("prenom", "")),
                "Nom": c.get("Nom", c.get("nom", "Non détecté")),
                "Email": c.get("Email", c.get("email", "")),
                "Téléphone": c.get("Téléphone", c.get("telephone", "")),
                "Profil": c.get("Profil", c.get("profil", "")),
                "Compétences": c.get("Compétences", c.get("competences", "")),
                "Diplômes": c.get("Diplômes", c.get("diplomes", ""))
            })
        
        df = pd.DataFrame(data)
        os.makedirs(os.path.dirname(chemin_final), exist_ok=True)
        
        NOM_ONGLET = "Candidats SONASID"
        df.to_excel(chemin_final, index=False, sheet_name=NOM_ONGLET)
        
        wb = load_workbook(chemin_final)
        ws = wb[NOM_ONGLET]
        
        VERT_FONCE = "1B4D3E"
        VERT_CLAIR_LIGNE = "F4F9F6"
        BLANC = "FFFFFF"
        
        font_header = Font(name="Segoe UI", size=11, bold=True, color=BLANC)
        fill_header = PatternFill(start_color=VERT_FONCE, end_color=VERT_FONCE, fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        font_data = Font(name="Segoe UI", size=10, color="000000")
        fill_even = PatternFill(start_color=VERT_CLAIR_LIGNE, end_color=VERT_CLAIR_LIGNE, fill_type="solid")
        border_thin = Border(
            left=Side(style='thin', color="E0E0E0"), right=Side(style='thin', color="E0E0E0"),
            top=Side(style='thin', color="E0E0E0"), bottom=Side(style='thin', color="E0E0E0")
        )
        
        ws.row_dimensions[1].height = 28
        
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = alignment_center
            cell.border = border_thin

        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 50
            is_even = (row_num % 2 == 0)
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = font_data
                cell.border = border_thin
                cell.alignment = alignment_left
                if is_even:
                    cell.fill = fill_even

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter in ['E', 'F', 'G']:
                ws.column_dimensions[col_letter].width = 45
            else:
                max_len = 0
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

        wb.save(chemin_final)